from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from django.db.models import Sum, Q, Count
from django.http import HttpResponse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import csv
import logging

logger = logging.getLogger(__name__)


class ExpensePagination(PageNumberPagination):
    page_size = 200
    page_size_query_param = 'page_size'
    max_page_size = 1000

from .models import BankStatement, Transaction, DecryptionKey, Category, Expense
from .serializers import (
    BankStatementSerializer,
    BankStatementUploadSerializer,
    TransactionSerializer,
    TransactionListSerializer,
    DecryptionKeySerializer,
    CategorySerializer,
    ExpenseSerializer,
    ExpenseListSerializer,
    ExpenseUpdateSerializer
)
from .services import BankStatementProcessorService


class BankStatementViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing bank statements.
    Supports uploading, processing, and viewing bank statements.
    """
    serializer_class = BankStatementSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'is_encrypted']
    search_fields = ['original_filename']
    ordering_fields = ['created_at', 'processed_at']
    ordering = ['-created_at']

    def get_queryset(self):
        """Filter bank statements by current user"""
        return BankStatement.objects.filter(user=self.request.user)

    @action(detail=False, methods=['post'])
    def upload(self, request):
        """
        Upload a new bank statement PDF.
        The file will be queued for processing.
        """
        # Debug logging
        logger.info(f"Upload request - Content-Type: {request.content_type}")
        logger.info(f"Upload request - Parser: {request.accepted_renderer}")
        logger.info(f"Upload request - Data keys: {request.data.keys()}")
        logger.info(f"Upload request - FILES: {request.FILES}")

        serializer = BankStatementUploadSerializer(data=request.data)
        if serializer.is_valid():
            file = serializer.validated_data['file']

            # Create bank statement record
            bank_statement = BankStatement.objects.create(
                user=request.user,
                file=file,
                original_filename=file.name,
                status='pending'
            )

            return Response({
                'id': bank_statement.id,
                'message': 'Bank statement uploaded successfully',
                'status': 'pending'
            }, status=status.HTTP_201_CREATED)

        logger.error(f"Upload validation errors: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def process(self, request, pk=None):
        """
        Process a bank statement to extract transactions.
        """
        bank_statement = self.get_object()

        if bank_statement.status == 'processing':
            return Response({
                'error': 'Bank statement is already being processed'
            }, status=status.HTTP_400_BAD_REQUEST)

        if bank_statement.status == 'completed':
            return Response({
                'error': 'Bank statement has already been processed'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Process the statement
        processor = BankStatementProcessorService(bank_statement)
        success = processor.process_statement()

        if success:
            serializer = self.get_serializer(bank_statement)
            return Response({
                'message': 'Bank statement processed successfully',
                'data': serializer.data
            }, status=status.HTTP_200_OK)
        else:
            return Response({
                'error': 'Failed to process bank statement',
                'message': bank_statement.error_message
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'])
    def transactions(self, request, pk=None):
        """
        Get all transactions for a specific bank statement.
        """
        bank_statement = self.get_object()
        transactions = bank_statement.transactions.all()

        serializer = TransactionListSerializer(transactions, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """
        Get statistics about bank statements and transactions.
        """
        user_statements = self.get_queryset()

        stats = {
            'total_statements': user_statements.count(),
            'pending': user_statements.filter(status='pending').count(),
            'processing': user_statements.filter(status='processing').count(),
            'completed': user_statements.filter(status='completed').count(),
            'failed': user_statements.filter(status='failed').count(),
        }

        return Response(stats)


class TransactionViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for viewing transactions.
    Transactions are created automatically when processing bank statements.
    """
    serializer_class = TransactionSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['date', 'bank_statement']
    search_fields = ['transaction_id_ref', 'cheque_ref_no']
    ordering_fields = ['date', 'created_at', 'credit', 'debit', 'balance']
    ordering = ['-date', '-created_at']

    def get_queryset(self):
        """Filter transactions by current user"""
        return Transaction.objects.filter(user=self.request.user)

    def get_serializer_class(self):
        """Use simplified serializer for list view"""
        if self.action == 'list':
            return TransactionListSerializer
        return TransactionSerializer

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """
        Get transaction summary (total credits, debits, etc.)
        Supports optional date filtering via query params:
        - start_date: YYYY-MM-DD
        - end_date: YYYY-MM-DD
        """
        queryset = self.get_queryset()

        # Apply date filters if provided
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)

        # Calculate summary
        summary = queryset.aggregate(
            total_credit=Sum('credit'),
            total_debit=Sum('debit'),
            transaction_count=Count('id')
        )

        # Get latest balance
        latest_transaction = queryset.order_by('-date', '-created_at').first()
        summary['latest_balance'] = latest_transaction.balance if latest_transaction else 0

        return Response(summary)

    @action(detail=False, methods=['get'])
    def by_month(self, request):
        """
        Get transactions grouped by month.
        Requires year parameter (e.g., ?year=2024)
        """
        year = request.query_params.get('year')
        if not year:
            return Response({
                'error': 'Year parameter is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            year = int(year)
        except ValueError:
            return Response({
                'error': 'Invalid year format'
            }, status=status.HTTP_400_BAD_REQUEST)

        queryset = self.get_queryset().filter(date__year=year)

        # Group by month
        monthly_data = []
        for month in range(1, 13):
            month_transactions = queryset.filter(date__month=month)
            month_summary = month_transactions.aggregate(
                total_credit=Sum('credit'),
                total_debit=Sum('debit'),
                transaction_count=Count('id')
            )
            month_summary['month'] = month
            month_summary['month_name'] = datetime(year, month, 1).strftime('%B')
            monthly_data.append(month_summary)

        return Response(monthly_data)


class DecryptionKeyViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing PDF decryption keys.
    Only accessible by admin users.
    """
    queryset = DecryptionKey.objects.all()
    serializer_class = DecryptionKeySerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['is_active']
    ordering = ['created_at']

    def get_queryset(self):
        """Only admins can view decryption keys"""
        if self.request.user.is_staff:
            return DecryptionKey.objects.all()
        return DecryptionKey.objects.none()


class CategoryViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing expense categories.
    Users can view default categories + their custom categories.
    """
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['is_default']
    search_fields = ['name']
    ordering = ['name']

    def get_queryset(self):
        """Get default categories + user's custom categories"""
        return Category.objects.filter(
            Q(is_default=True) | Q(user=self.request.user)
        )

    @action(detail=False, methods=['get'])
    def all(self, request):
        """Get all categories without statistics (simple list)"""
        queryset = self.get_queryset()

        categories = []
        for category in queryset:
            categories.append({
                'id': category.id,
                'name': category.name,
                'color': category.color,
                'created_at': category.created_at.isoformat()
            })

        return Response(categories)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get category statistics with expense counts"""
        queryset = self.get_queryset()

        category_stats = list(
            queryset.annotate(
                expense_count=Count('expense', filter=Q(expense__user=request.user)),
                total_amount=Sum('expense__amount', filter=Q(expense__user=request.user)),
            ).values('id', 'name', 'icon', 'color', 'expense_count', 'total_amount')
        )
        for stat in category_stats:
            stat['total_amount'] = float(stat['total_amount'] or 0)

        return Response(category_stats)


class ExpenseViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing expenses.
    Expenses are auto-created from transactions but can be manually categorized.
    """
    serializer_class = ExpenseSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = ExpensePagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'date', 'is_recurring', 'is_auto_categorized']
    search_fields = ['description', 'notes']
    ordering_fields = ['date', 'amount', 'created_at']
    ordering = ['-date', '-created_at']

    def get_queryset(self):
        """Filter expenses by current user"""
        queryset = Expense.objects.filter(user=self.request.user)

        # Optional amount range filtering
        min_amount = self.request.query_params.get('min_amount')
        max_amount = self.request.query_params.get('max_amount')

        if min_amount:
            queryset = queryset.filter(amount__gte=min_amount)
        if max_amount:
            queryset = queryset.filter(amount__lte=max_amount)

        return queryset

    def get_serializer_class(self):
        """Use different serializers for different actions"""
        if self.action == 'list':
            return ExpenseListSerializer
        elif self.action in ['update', 'partial_update']:
            return ExpenseUpdateSerializer
        return ExpenseSerializer

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """
        Get expense summary (total amount, count, etc.)
        Supports optional date filtering via query params:
        - start_date: YYYY-MM-DD
        - end_date: YYYY-MM-DD
        - category: category ID
        """
        queryset = self.get_queryset()

        # Apply date filters if provided
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        category_id = request.query_params.get('category')

        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)
        if category_id:
            queryset = queryset.filter(category_id=category_id)

        # Calculate summary
        summary = queryset.aggregate(
            total_amount=Sum('amount'),
            expense_count=Count('id'),
            avg_amount=Sum('amount') / Count('id') if queryset.exists() else 0
        )

        # Get category breakdown
        category_breakdown = queryset.values(
            'category__name', 'category__icon', 'category__color'
        ).annotate(
            total=Sum('amount'),
            count=Count('id')
        ).order_by('-total')

        summary['category_breakdown'] = list(category_breakdown)
        summary['uncategorized_count'] = queryset.filter(category__isnull=True).count()

        return Response(summary)

    @action(detail=False, methods=['get'])
    def by_month(self, request):
        """
        Get expenses grouped by month.
        Requires year parameter (e.g., ?year=2025)
        """
        year = request.query_params.get('year')
        if not year:
            return Response({
                'error': 'Year parameter is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            year = int(year)
        except ValueError:
            return Response({
                'error': 'Invalid year format'
            }, status=status.HTTP_400_BAD_REQUEST)

        queryset = self.get_queryset().filter(date__year=year)

        # Group by month
        monthly_data = []
        for month in range(1, 13):
            month_expenses = queryset.filter(date__month=month)
            month_summary = month_expenses.aggregate(
                total_amount=Sum('amount'),
                expense_count=Count('id')
            )
            month_summary['month'] = month
            month_summary['month_name'] = datetime(year, month, 1).strftime('%B')
            monthly_data.append(month_summary)

        return Response(monthly_data)

    @action(detail=False, methods=['get'])
    def by_category(self, request):
        """
        Get expenses grouped by category.
        Supports optional date filtering via query params.
        """
        queryset = self.get_queryset()

        # Apply date filters if provided
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)

        # Group by category
        category_data = queryset.values(
            'category__id', 'category__name', 'category__icon', 'category__color'
        ).annotate(
            total_amount=Sum('amount'),
            expense_count=Count('id')
        ).order_by('-total_amount')

        return Response(list(category_data))

    @action(detail=False, methods=['get'])
    def recurring(self, request):
        """Get all recurring expenses"""
        queryset = self.get_queryset().filter(is_recurring=True)
        serializer = ExpenseListSerializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def uncategorized(self, request):
        """Get all uncategorized expenses"""
        queryset = self.get_queryset().filter(category__isnull=True)
        serializer = ExpenseListSerializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def export(self, request):
        """
        Export expenses to Excel or CSV file.
        Query param: format=excel (default) or format=csv
        Supports same filtering as list endpoint: category, date, min_amount, max_amount, etc.
        """
        # Get export format from query params
        export_format = request.query_params.get('format', 'excel').lower()

        if export_format not in ['excel', 'csv']:
            return Response({
                'error': 'Invalid format. Use format=excel or format=csv'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Get filtered queryset (respects all query parameters except 'format')
        # Create a mutable copy of query params without 'format'
        filter_params = request.query_params.copy()
        filter_params.pop('format', None)

        # Temporarily replace query_params for filtering
        original_query_params = request._request.GET
        request._request.GET = filter_params
        queryset = self.filter_queryset(self.get_queryset())
        request._request.GET = original_query_params

        if export_format == 'excel':
            return self._export_excel(request, queryset)
        else:
            return self._export_csv(request, queryset)

    def _export_csv(self, request, queryset):
        """Export expenses to CSV file"""
        # Optimize queryset with select_related
        queryset = queryset.select_related('category', 'transaction')

        # Calculate total upfront to avoid double aggregation
        queryset_count = queryset.count()
        total_amount = queryset.aggregate(total=Sum('amount'))['total'] or 0
        logger.info(f"CSV Export: Processing {queryset_count} expenses, total: {total_amount}")

        # Create HTTP response with UTF-8 BOM for Excel compatibility
        response = HttpResponse(content_type='text/csv; charset=utf-8')

        # Generate filename with date range if available
        start_date = request.query_params.get('start_date', 'all')
        end_date = request.query_params.get('end_date', 'all')
        filename = f'expenses_{start_date}_to_{end_date}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Add UTF-8 BOM for Excel to recognize UTF-8 encoding
        response.write('\ufeff')

        # Create CSV writer
        writer = csv.writer(response)

        # Write headers
        writer.writerow([
            'Date',
            'Description',
            'Amount',
            'Category',
            'Notes',
            'Is Recurring',
            'Tags',
            'Transaction Reference',
            'Auto Categorized'
        ])

        # Write data
        rows_written = 0
        for expense in queryset:
            rows_written += 1
            writer.writerow([
                expense.date.strftime('%Y-%m-%d'),
                expense.description,
                float(expense.amount),
                expense.category.name if expense.category else 'Uncategorized',
                expense.notes or '',
                'Yes' if expense.is_recurring else 'No',
                ', '.join(expense.tags) if expense.tags else '',
                expense.transaction.transaction_id_ref,
                'Yes' if expense.is_auto_categorized else 'No'
            ])

        # Write summary row (use pre-calculated total_amount)
        writer.writerow([])  # Empty row
        writer.writerow(['TOTAL', '', float(total_amount)])

        logger.info(f"CSV Export: Wrote {rows_written} expense rows")
        return response

    def _export_excel(self, request, queryset):
        """Export expenses to Excel file"""
        # Optimize queryset with select_related
        queryset = queryset.select_related('category', 'transaction')

        # Calculate total upfront to avoid double aggregation
        queryset_count = queryset.count()
        total_amount = queryset.aggregate(total=Sum('amount'))['total'] or 0
        logger.info(f"Excel Export: Processing {queryset_count} expenses, total: {total_amount}")

        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Expenses'

        # Define styles
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = [
            'Date',
            'Description',
            'Amount (₹)',
            'Category',
            'Notes',
            'Is Recurring',
            'Tags',
            'Transaction Reference',
            'Auto Categorized'
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Write data
        rows_written = 0
        for row_num, expense in enumerate(queryset, 2):
            rows_written += 1
            ws.cell(row=row_num, column=1, value=expense.date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=2, value=expense.description)
            ws.cell(row=row_num, column=3, value=float(expense.amount))
            ws.cell(row=row_num, column=4, value=expense.category.name if expense.category else 'Uncategorized')
            ws.cell(row=row_num, column=5, value=expense.notes or '')
            ws.cell(row=row_num, column=6, value='Yes' if expense.is_recurring else 'No')
            ws.cell(row=row_num, column=7, value=', '.join(expense.tags) if expense.tags else '')
            ws.cell(row=row_num, column=8, value=expense.transaction.transaction_id_ref)
            ws.cell(row=row_num, column=9, value='Yes' if expense.is_auto_categorized else 'No')

            # Apply border to all cells
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).border = border

        # Set column widths (optimized - avoid iterating all cells)
        column_widths = {
            'A': 12,  # Date
            'B': 35,  # Description
            'C': 15,  # Amount
            'D': 20,  # Category
            'E': 30,  # Notes
            'F': 14,  # Is Recurring
            'G': 25,  # Tags
            'H': 20,  # Transaction Reference
            'I': 18,  # Auto Categorized
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Add summary row (use pre-calculated total_amount)
        summary_row = queryset_count + 3
        ws.cell(row=summary_row, column=1, value='TOTAL').font = Font(bold=True)
        ws.cell(row=summary_row, column=3, value=float(total_amount)).font = Font(bold=True)

        # Number format for amount column
        for row in range(2, queryset_count + 2):
            ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=summary_row, column=3).number_format = '#,##0.00'

        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Generate filename with date range if available
        start_date = request.query_params.get('start_date', 'all')
        end_date = request.query_params.get('end_date', 'all')
        filename = f'expenses_{start_date}_to_{end_date}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Save workbook to response
        wb.save(response)

        logger.info(f"Excel Export: Wrote {rows_written} expense rows")
        return response

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """
        Legacy endpoint for Excel export only.
        Use /export/?format=excel or /export/?format=csv instead.
        """
        queryset = self.filter_queryset(self.get_queryset())
        return self._export_excel(request, queryset)

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """
        CSV export endpoint.
        Use /export/?format=csv instead (recommended).
        """
        queryset = self.filter_queryset(self.get_queryset())
        return self._export_csv(request, queryset)
